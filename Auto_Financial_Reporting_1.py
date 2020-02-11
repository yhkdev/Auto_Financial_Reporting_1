import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string as get_col_idx

from Rule import Rule


# Set up - Import files (Identify main XL and 3 other XL to copy from)

# paste_wb = openpyxl.load_workbook(filename='PasteTo.xlsx', data_only=True)  # 'data_only' makes sure you only get data, Not formula in the cells
# paste_ws = paste_wb['Sheet1']



def _copy_paste(copy_ws, paste_ws, copyfrom, pasteto):
    """ Copy & Paste single cell or multiple cells of values to another sheet
    :param copy_ws: Worksheet (NOT Workbook) to copy from
    :param paste_ws: Worksheet (NOT Workbook) to paste to
    :param copyfrom: range (ex: "B2" or "B2:D4") to copy from
    :param pasteto: range to paste to
    """

    if ":" in copyfrom:
        # 2D Matrix Copy & Paste - "B2:D4" -> "C2:E4" in other WS
        for copy_row, paste_row in zip(copy_ws[copyfrom], paste_ws[pasteto]):
            for copy_cell, paste_cell in zip(copy_row, paste_row):
                paste_cell.value = copy_cell.value
    else:
        # Single Cell Copy & Paste - 'B2' -> 'B2' in other WS
        paste_ws[pasteto].value = copy_ws[copyfrom].value

def copy_paste(copy_ws, paste_ws, entries):
    """ _copy_paste() on copy/paste range in given dictionary
    :param copy_paste_dict: Dictionary of {copy_range : paste_range}
    """
    for copyfrom, pasteto in entries:
        _copy_paste(copy_ws, paste_ws, copyfrom, pasteto)



# ----- Test -----
test_wb = Workbook()
os.chdir(os.getcwd() + "/local_test")  # Change working directory to folder with the excel files
copy_wb = openpyxl.load_workbook(filename='CopyFrom.xlsx', data_only=True)
copy_ws = copy_wb['Sheet1']

def test_copy_paste(copy_ws, test_ws, copy_range, paste_range):
    """ Test function for copy_paste() and _copy_paste()"""
    if ":" in copy_range:
        for copy_row, paste_row in zip(copy_ws[copy_range], test_ws[paste_range]):
            for copy_cell, paste_cell in zip(copy_row, paste_row):
                assert paste_cell.value == copy_cell.value
    else:
        assert test_ws[paste_range].value == copy_ws[copy_range].value


# _copy_paste()  --  Single Cell input
test_ws = test_wb.create_sheet()
copy_loc, paste_loc = "B2", "C2"
_copy_paste(copy_ws, test_ws, copy_loc, paste_loc)
test_copy_paste(copy_ws, test_ws, copy_loc, paste_loc)

# _copy_paste()  --  Multi-Cell input
test_ws = test_wb.create_sheet()
copy_range, paste_range = "B2:D4", "C2:E4"
_copy_paste(copy_ws, test_ws, copy_range, paste_range)
test_copy_paste(copy_ws, test_ws, copy_range, paste_range)

# copy_paste()  --  both Single cell & Multi-Cell input
test_ws = test_wb.create_sheet()
entries = [["B2", "B2"], ["B3", "B4"], ["C2:C4", "D2:D4"]]
copy_paste(copy_ws, test_ws, entries)
for copy_range, paste_range in entries:
    test_copy_paste(copy_ws, test_ws, copy_range, paste_range)

# Rule class integration
test_ws = test_wb.create_sheet()
rule = Rule("Rule Name", "Rule Description")
rule.add_entry("B2", "B2")
rule.add_entry("C2:C4", "D2:D4")
copy_paste(copy_ws, test_ws, rule.entries)
for copy_range, paste_range in rule.entries:
    test_copy_paste(copy_ws, test_ws, copy_range, paste_range)



# Check if XL format of reports are still the same (ex: no extra column item)
col_items = ['Fundraising', 'G&A', 'Programs', 'Arconic Grant', 'Disaster Relief (Restricted Fd)', 'Total Programs',
             'Restricted Funds', 'TOTAL']  # Cell[B5:I5]
row_items = ['Leadership Contributions', 'Individual Contrilbutions', 'Corporate Contributions']  # Cell[A6:A9]
#
# def check_wb(wb, col_items, row_items):
#     ws = A_wb['Sheet1']
#
#     try:  # Check Columns
#         for i in range(2, len(col_items)):
#             assert ws.cell(row=5, column=i).value == col_items[i-2]
#     except:
#         raise Exception("Columns of the QBO report sheet has changed!")
#
#     try:  # Check Rows
#         for i in range(len(row_items)):
#             assert ws.cell(row=i+7, column=1).value == row_items[i]
#     except:
#         raise Exception("Rows of the QBO report sheet has changed!")
# # check_wb(A_wb, col_items, row_items)
#
#
# # If pass: Run the default hardcoded copy-paste algo
#
# copy_start, copy_end = "I7", "I9"
# paste_start, paste_end = "C9", "C11"
#
# copy_loc = ("I7","I9")
# paste_loc = ("C9","C11")
#
# def copy_paste_cells(copy_wb, main_wb):
#     pass
#
# ws = A_wb['Sheet1']
# main_ws = main_wb['Sheet1']
#
# for idx in range(3):
#     print(ws.cell(row=idx+int(copy_loc[0][1]), column=get_col_idx(copy_loc[0][0])).value, end=" ")
#     print(main_ws.cell(row=idx+int(paste_loc[0][1]), column=get_col_idx(paste_loc[0][0])).value)
#
#     main_ws.cell(row=idx+int(paste_loc[0][1]), column=get_col_idx(paste_loc[0][0])).value = ws.cell(row=idx+int(copy_loc[0][1]), column=get_col_idx(copy_loc[0][0])).value
#
#     print(ws.cell(row=idx + int(copy_loc[0][1]), column=get_col_idx(copy_loc[0][0])).value, end=" ")
#     print(main_ws.cell(row=idx + int(paste_loc[0][1]), column=get_col_idx(paste_loc[0][0])).value)



# Else: Identify each col/row items and find cells to copy/paste from. Then run copy/paste algo

# Save file?
